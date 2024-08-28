#!/usr/bin/python3
# -*- coding: UTF-8 -*-

#------------------------------------------------------------------------------------------------------------#
#Usage: ./jiratool.py username password
#------------------------------------------------------------------------------------------------------------#

from jira import JIRA
import re
import os
import sys
import datetime
import time
from openpyxl import Workbook

#------------------------------------------------------------------------------------------------------------#
#Logon to jira system with windows username and password
#------------------------------------------------------------------------------------------------------------#
#jira = JIRA('https://plf-jira.rnd.ki.sw.ericsson.se',basic_auth=(sys.argv[1], sys.argv[2]))
jira = JIRA('https://eteamproject.internal.ericsson.com',basic_auth=(sys.argv[1], sys.argv[2]))
#------------------------------------------------------------------------------------------------------------#

#------------------------------------------------------------------------------------------------------------#
#Search all Epic belong to CCRC, and Search all User Story from the jira system
#------------------------------------------------------------------------------------------------------------#
#req_issues = jira.search_issues('project = "5G UDM & Policy" AND (issuetype = "Product Requirement" OR issuetype = "Commercial Requirement" AND cf[12310] = "5G Core UDM & Policy" OR issuetype = "Internal Requirement" AND cf[12310] = "5G Core UDM & Policy") ORDER BY Rank ASC', maxResults=None)
epic_issues = jira.search_issues('project = UDM5GP AND (issuetype = Epic AND "FP Product" ~ nrf OR "FP Product" ~ nssf) ORDER BY Rank ASC', maxResults=None)
us_issues = jira.search_issues('project = "5G UDM & Policy" AND (issuetype = "User Story") ORDER BY Rank ASC', maxResults=None)
#------------------------------------------------------------------------------------------------------------#

#------------------------------------------------------------------------------------------------------------#
# Find out the requirement info from link in Epic issues, output mapping: Requirement <-> Epic
#------------------------------------------------------------------------------------------------------------#
ll1 = []
for epic_issue in epic_issues:
    for link in epic_issue.fields.issuelinks:
        if hasattr(link, "outwardIssue"):
            req_issue = link.outwardIssue
        if hasattr(link, "inwardIssue"):
            req_issue = link.inwardIssue
        if (req_issue.fields.issuetype.name == "Product Requirement") or (req_issue.fields.issuetype.name == "Commercial Requirement") or (req_issue.fields.issuetype.name == "Internal Requirement"):
            l1 = []
            l1.append(req_issue.key)
            l1.append(req_issue.fields.summary)
            l1.append(epic_issue.key)
            l1.append(epic_issue.fields.summary)
            l1.append(epic_issue.fields.customfield_18719)
            l1.append(epic_issue.fields.status)
            ll1.append(l1)
    if epic_issue.fields.issuelinks == []:
        l1 = []
        l1.append("None")
        l1.append("None")
        l1.append(epic_issue.key)
        l1.append(epic_issue.fields.summary)
        l1.append(epic_issue.fields.customfield_18719)
        l1.append(epic_issue.fields.status)
        ll1.append(l1)
#------------------------------------------------------------------------------------------------------------#

#------------------------------------------------------------------------------------------------------------#
# Find out the Epic and Test Case info from link in User Story issues, output mapping: Epic <-> US <-> TC
#------------------------------------------------------------------------------------------------------------#
ll2 = []
for us_issue in us_issues:
    epic_list = us_issue.raw['fields']['customfield_11911']
    if epic_list != None:
        tc_flag = False
        for tc_list in us_issue.raw['fields']['customfield_18913']:
            for k, v in tc_list.items():
                if (k in ("okJql","nokJql","notRunJql","unknownJql")) and len(v) > 14:
                    strlist = v[13:len(v)-1].split(',')
                    for j in strlist:
                        tc_flag = True
                        l2 = []
                        l2.append(epic_list)
                        l2.append(us_issue.key)
                        l2.append(us_issue.fields.summary)
                        l2.append(us_issue.fields.status)
                        l2.append(j.lstrip().rstrip())
                        ll2.append(l2)
        if not tc_flag:
            l2 = []
            l2.append(epic_list)
            l2.append(us_issue.key)
            l2.append(us_issue.fields.summary)
            l2.append(us_issue.fields.status)
            l2.append("None")
            ll2.append(l2)
#------------------------------------------------------------------------------------------------------------#

#------------------------------------------------------------------------------------------------------------#
# Handle two List (Requirement <-> Epic) and (Epic <-> US <-> TC), output is : Requirement <-> Epic <-> US <-> TC
#------------------------------------------------------------------------------------------------------------#
ll3 = []
for i in ll1:
    flag = False
    for k in ll2:
        if i[2] == k[0]:
            flag = True
            l3 = []
            for j in i:
                l3.append(j)
            for j in k:
                if j == k[0]:
                    continue
                l3.append(j)
            ll3.append(l3)
    if not flag:
        l3 = []        
        for j in i:
            l3.append(j)
        for k in range(4):
            l3.append("None")
        ll3.append(l3)
#------------------------------------------------------------------------------------------------------------#

#------------------------------------------------------------------------------------------------------------#
# Output to excel file according to List: Requirement <-> Epic <-> US <-> TC
#------------------------------------------------------------------------------------------------------------#
wb = Workbook()
ws = wb.active

ws['A1'] = "Requirement Key"
ws['B1'] = "Requirement Summary"
ws['C1'] = "Epic Key"
ws['D1'] = "Epic Summary"
ws['E1'] = "Network Function"
ws['F1'] = "Epic Status"
ws['G1'] = "User Story Key"
ws['H1'] = "User Sotry Summary"
ws['I1'] = "User Sotry Status"
ws['J1'] = "Test Case"

k = 1
for i in ll3:
    k = k + 1
    ws['A' + str(k)] = str(i[0])
    ws['B' + str(k)] = str(i[1])
    ws['C' + str(k)] = str(i[2])
    ws['D' + str(k)] = str(i[3])
    ws['E' + str(k)] = str(i[4])
    ws['F' + str(k)] = str(i[5])
    ws['G' + str(k)] = str(i[6])
    ws['H' + str(k)] = str(i[7])
    ws['I' + str(k)] = str(i[8])
    ws['J' + str(k)] = str(i[9])

wb.save("jira-"+ str(datetime.datetime.now().date()) + "_" + str(time.strftime("%H-%M-%S")) + ".xlsx")
print("Great! " + "jira-"+ str(datetime.datetime.now().date()) + "_" + str(time.strftime("%H-%M-%S")) + ".xlsx " + "output successfully.")
#------------------------------------------------------------------------------------------------------------#
#------------------------------------------------------------------------------------------------------------#